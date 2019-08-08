import rqdatac
from rqdatac import *
import pandas as pd
from pandas import *
import numpy as np
from numpy import *
import openpyxl
from openpyxl import *
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook,InvalidFileException

import sys
import rqdatac_bond
from rqdatac_bond import *
import re

bond_list = []

rqdatac.init(uri='tcp://rice:rice@192.168.10.106:16010')

def translate_id():
    delete_SH= re.compile(r'.SH')
    delete_SZ= re.compile(r'.SZ')
    index = 0;
    for bond in bond_list:
        pre_id = bond["order_book_id"]
        if pre_id[-3:]==".SH":
            new_id = delete_SH.sub('',pre_id)
            new_id +='.XSHG'
            bond_list[index]["order_book_id"] = new_id
        elif pre_id[-3:]==".SZ":
            new_id = delete_SZ.sub('',pre_id)
            new_id +='.XSHE'
            bond_list[index]["order_book_id"] = new_id
        index+=1           



        
    

def main(argv):
    
    workbook_name  = argv[1]
    wb = load_workbook(workbook_name)
    sheet = wb.get_sheet_by_name(wb.sheetnames[0])
    date_value = sheet['C1'].value
    date_value = date_value.strftime('%Y/%m/%d')

    total_equity = 10000000000 #总权益为100亿
    cash = 0
    position_value_sum = 0
    for row in sheet.iter_rows(min_row = 2):
        temp_list = {}   
        order_book_id = row[2].internal_value
        position_weight = row[4].internal_value
        position_value = position_weight * total_equity
        dataframe = bond.get_price(order_book_id,'2019-04-30','2019-04-30','chinabond',details=True)
        position_value_sum += position_value
        if dataframe is None:
            cash +=position_value
        else: 
            dirty_price_eod = dataframe.dirty_price_eod.values[0]
            quantity = round(position_value/dirty_price_eod)
            temp_list = {'date':date_value, 'order_book_id':order_book_id, 'quantity' :quantity, 'position_weight' :position_weight,'position_value': position_value,'dirty_price_eod': dirty_price_eod}
            bond_list.append(temp_list)  
    print('cash:',cash)
    #translate_id()

    output_wb=Workbook()#创建一个工作簿
    ws=output_wb.active#获取工作的激活工作表
    ws['A1']="date"
    ws['B1']="order_book_id"
    ws['C1']="quantity"
    ws['D1']="position_weight"
    ws['E1']="position_value"
    ws['F1']="dirty_price_eod"
           
    count = 2
    for i in bond_list:
        ws.cell(row= count , column = 1) .value = i["date"]
        ws.cell(row= count , column = 2) .value = i["order_book_id"]
        ws.cell(row= count , column = 3) .value = i["quantity"]
        ws.cell(row= count , column = 4) .value = i["position_weight"]
        ws.cell(row= count , column = 5) .value = i["position_value"]
        ws.cell(row= count , column = 6) .value = i["dirty_price_eod"]
        count += 1

    ws.cell(row= count , column = 1) .value = date_value
    ws.cell(row= count , column = 2) .value = "cash"
    ws.cell(row= count , column = 3) .value = cash+(total_equity-position_value_sum)

    output_wb.save('output.xlsx')
   

if __name__ == "__main__":
    main(argv=sys.argv)


#bond.get_price('100001.IB','2018-01-05','2012-01-06','chinabond',details=True)